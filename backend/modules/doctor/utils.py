"""
医生管理子系统 - 工具函数
Doctor Management - Utility Functions

包含常用的业务逻辑函数
"""
from datetime import datetime, date, timedelta
from flask import request, jsonify
from extensions import db
from models import Doctor, DoctorSchedule, DoctorPerformance, Appointment, MedicalRecord
from typing import Optional, List, Dict
from functools import wraps
from marshmallow import ValidationError
import json


# ============= 排班相关工具函数 =============

def check_schedule_conflict(doctor_id: int, schedule_date: date, start_time: str, 
                           end_time: str, exclude_schedule_id: Optional[int] = None) -> bool:
    """
    检查排班冲突
    
    Args:
        doctor_id: 医生ID
        schedule_date: 排班日期
        start_time: 开始时间 (HH:MM格式)
        end_time: 结束时间 (HH:MM格式)
        exclude_schedule_id: 排除的排班ID(用于更新时)
    
    Returns:
        bool: True表示有冲突，False表示无冲突
    """
    query = DoctorSchedule.query.filter_by(
        doctor_id=doctor_id,
        date=schedule_date
    )
    
    # 排除指定的排班记录（用于更新场景）
    if exclude_schedule_id:
        query = query.filter(DoctorSchedule.id != exclude_schedule_id)
    
    # 检查时间冲突
    conflicts = query.filter(
        db.or_(
            # 新排班的开始时间在已有排班的时间范围内
            db.and_(
                DoctorSchedule.start_time <= start_time,
                DoctorSchedule.end_time > start_time
            ),
            # 新排班的结束时间在已有排班的时间范围内
            db.and_(
                DoctorSchedule.start_time < end_time,
                DoctorSchedule.end_time >= end_time
            ),
            # 新排班完全包含已有排班
            db.and_(
                DoctorSchedule.start_time >= start_time,
                DoctorSchedule.end_time <= end_time
            )
        )
    ).first()
    
    return conflicts is not None


def check_leave_conflict(doctor_id: int, start_date: date, end_date: date,
                        exclude_leave_id: Optional[int] = None) -> bool:
    """
    检查请假冲突
    
    Args:
        doctor_id: 医生ID
        start_date: 请假开始日期
        end_date: 请假结束日期
        exclude_leave_id: 排除的请假ID(用于更新时)
    
    Returns:
        bool: True表示有冲突，False表示无冲突
    """
    from modules.doctor.models_extended import DoctorLeave
    
    query = DoctorLeave.query.filter_by(doctor_id=doctor_id)
    
    # 只检查已批准的请假
    query = query.filter(DoctorLeave.status == 'approved')
    
    if exclude_leave_id:
        query = query.filter(DoctorLeave.id != exclude_leave_id)
    
    # 检查日期冲突
    conflicts = query.filter(
        db.or_(
            db.and_(
                DoctorLeave.start_date <= start_date,
                DoctorLeave.end_date >= start_date
            ),
            db.and_(
                DoctorLeave.start_date <= end_date,
                DoctorLeave.end_date >= end_date
            ),
            db.and_(
                DoctorLeave.start_date >= start_date,
                DoctorLeave.end_date <= end_date
            )
        )
    ).first()
    
    return conflicts is not None


def calculate_leave_days(start_date: date, end_date: date) -> int:
    """
    计算请假天数（包含开始和结束日期）
    
    Args:
        start_date: 开始日期
        end_date: 结束日期
    
    Returns:
        int: 请假天数
    """
    return (end_date - start_date).days + 1


def get_available_schedules(department: Optional[str] = None, 
                           schedule_date: Optional[date] = None,
                           shift: Optional[str] = None) -> List[Dict]:
    """
    获取可用的排班列表
    
    Args:
        department: 科室
        schedule_date: 排班日期
        shift: 班次
    
    Returns:
        List[Dict]: 可用排班列表
    """
    query = DoctorSchedule.query.join(Doctor).filter(
        DoctorSchedule.status == 'available',
        Doctor.status == 'active'
    )
    
    if department:
        query = query.filter(Doctor.department == department)
    
    if schedule_date:
        query = query.filter(DoctorSchedule.date == schedule_date)
    
    if shift:
        query = query.filter(DoctorSchedule.shift == shift)
    
    schedules = query.all()
    return [schedule.to_dict() for schedule in schedules]


# ============= 绩效统计工具函数 =============

def calculate_doctor_performance(doctor_id: int, year: int, month: int) -> Dict:
    """
    自动计算医生绩效数据
    
    Args:
        doctor_id: 医生ID
        year: 年份
        month: 月份
    
    Returns:
        Dict: 绩效数据
    """
    # 计算月份的开始和结束日期
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    # 统计接诊人数
    try:
        patient_count = Appointment.query.filter(
            Appointment.doctor_id == doctor_id,
            Appointment.appointment_date >= start_date,
            Appointment.appointment_date <= end_date,
            Appointment.status.in_(['completed', 'confirmed'])
        ).count()
    except:
        patient_count = 0
    
    # 统计排班次数
    try:
        schedule_count = DoctorSchedule.query.filter(
            DoctorSchedule.doctor_id == doctor_id,
            DoctorSchedule.date >= start_date,
            DoctorSchedule.date <= end_date
        ).count()
    except:
        schedule_count = 0
    
    # 统计病历数
    try:
        medical_record_count = MedicalRecord.query.filter(
            MedicalRecord.doctor_id == doctor_id,
            MedicalRecord.visit_date >= start_date,
            MedicalRecord.visit_date <= end_date
        ).count()
    except:
        medical_record_count = 0
    
    return {
        'patient_count': patient_count,
        'schedule_count': schedule_count,
        'medical_record_count': medical_record_count,
        'month_days': (end_date - start_date).days + 1
    }


# ============= 操作日志工具函数 =============

def log_operation(operation: str, resource: str, resource_id: Optional[int] = None,
                 resource_name: Optional[str] = None, details: Optional[Dict] = None,
                 user_id: Optional[int] = None, username: Optional[str] = None,
                 status: str = 'success', error_message: Optional[str] = None):
    """
    记录操作日志
    
    Args:
        operation: 操作类型 (create/update/delete/view/export)
        resource: 资源类型 (doctor/schedule/performance/qualification/leave)
        resource_id: 资源ID
        resource_name: 资源名称
        details: 操作详情
        user_id: 操作用户ID
        username: 操作用户名
        status: 状态 (success/failed)
        error_message: 错误信息
    """
    from modules.doctor.models_extended import OperationLog
    
    # 获取请求信息
    ip_address = request.remote_addr if request else None
    user_agent = request.headers.get('User-Agent', '')[:200] if request else None
    
    # 将详情转换为JSON字符串
    details_json = json.dumps(details, ensure_ascii=False) if details else None
    
    # 创建日志记录
    log = OperationLog(
        user_id=user_id,
        username=username,
        operation=operation,
        resource=resource,
        resource_id=resource_id,
        resource_name=resource_name,
        details=details_json,
        ip_address=ip_address,
        user_agent=user_agent,
        status=status,
        error_message=error_message
    )
    
    try:
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Failed to log operation: {str(e)}")


# ============= 通知工具函数 =============

def send_notification(user_id: int, title: str, content: str, 
                     notification_type: str, priority: str = 'normal',
                     related_resource: Optional[str] = None,
                     related_id: Optional[int] = None):
    """
    发送通知
    
    Args:
        user_id: 接收用户ID
        title: 通知标题
        content: 通知内容
        notification_type: 通知类型 (schedule/performance/system/leave/qualification)
        priority: 优先级 (low/normal/high/urgent)
        related_resource: 关联资源类型
        related_id: 关联资源ID
    """
    from modules.doctor.models_extended import Notification
    
    notification = Notification(
        user_id=user_id,
        title=title,
        content=content,
        type=notification_type,
        priority=priority,
        related_resource=related_resource,
        related_id=related_id
    )
    
    try:
        db.session.add(notification)
        db.session.commit()
        return notification.id
    except Exception as e:
        db.session.rollback()
        print(f"Failed to send notification: {str(e)}")
        return None


def send_qualification_expiry_alert(doctor_id: int, qualification_id: int, 
                                   expiry_date: date):
    """
    发送资质即将过期提醒
    
    Args:
        doctor_id: 医生ID
        qualification_id: 资质ID
        expiry_date: 过期日期
    """
    doctor = Doctor.query.get(doctor_id)
    if not doctor:
        return
    
    days_until_expiry = (expiry_date - date.today()).days
    
    title = f"资质证书即将过期提醒"
    content = f"您的资质证书将在{days_until_expiry}天后（{expiry_date}）过期，请及时更新。"
    
    # 假设医生有关联的用户ID（实际应从数据库关联获取）
    # 这里需要根据实际的User-Doctor关联关系来获取user_id
    # send_notification(user_id, title, content, 'qualification', 'high', 'qualification', qualification_id)


# ============= 数据导出工具函数 =============

def export_doctors_to_excel(doctors: List[Doctor], filename: str = 'doctors.xlsx'):
    """
    导出医生数据到Excel
    
    Args:
        doctors: 医生列表
        filename: 文件名
    
    Returns:
        BytesIO: Excel文件的二进制流
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "医生信息"
        
        # 设置表头
        headers = ['医生编号', '姓名', '性别', '年龄', '科室', '职称', '专长', 
                  '联系电话', '邮箱', '入职日期', '状态']
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入数据
        for doctor in doctors:
            ws.append([
                doctor.doctor_no,
                doctor.name,
                doctor.gender,
                doctor.age,
                doctor.department,
                doctor.title,
                doctor.specialty,
                doctor.phone,
                doctor.email,
                doctor.hire_date.isoformat() if doctor.hire_date else '',
                '在职' if doctor.status == 'active' else '离职'
            ])
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    except ImportError:
        raise Exception("需要安装openpyxl库：pip install openpyxl")


# ============= 数据验证工具函数 =============

def validate_doctor_data(data: Dict) -> tuple[bool, Optional[str]]:
    """
    验证医生数据
    
    Args:
        data: 医生数据字典
    
    Returns:
        tuple: (是否有效, 错误信息)
    """
    from modules.doctor.schemas import DoctorSchema
    
    schema = DoctorSchema()
    try:
        schema.load(data)
        return True, None
    except ValidationError as err:
        return False, str(err.messages)


def validate_schedule_data(data: Dict) -> tuple[bool, Optional[str]]:
    """
    验证排班数据
    
    Args:
        data: 排班数据字典
    
    Returns:
        tuple: (是否有效, 错误信息)
    """
    from modules.doctor.schemas import DoctorScheduleSchema
    
    schema = DoctorScheduleSchema()
    try:
        schema.load(data)
        return True, None
    except ValidationError as err:
        return False, str(err.messages)


def validate_performance_data(data: Dict) -> tuple[bool, Optional[str]]:
    """
    验证绩效数据
    
    Args:
        data: 绩效数据字典
    
    Returns:
        tuple: (是否有效, 错误信息)
    """
    from modules.doctor.schemas import DoctorPerformanceSchema
    
    schema = DoctorPerformanceSchema()
    try:
        schema.load(data)
        return True, None
    except ValidationError as err:
        return False, str(err.messages)


def validate_with_schema(schema_class):
    """
    数据验证装饰器
    
    Args:
        schema_class: Marshmallow Schema类
    
    Returns:
        装饰器函数
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            data = request.get_json()
            if not data:
                return jsonify({
                    'success': False,
                    'message': '请求数据不能为空',
                    'code': 'INVALID_DATA',
                    'data': None
                }), 400
            
            schema = schema_class()
            try:
                # 验证并加载数据
                validated_data = schema.load(data)
                # 将验证后的数据传递给路由函数
                request.validated_data = validated_data
                return f(*args, **kwargs)
            except ValidationError as err:
                return jsonify({
                    'success': False,
                    'message': '数据验证失败',
                    'code': 'VALIDATION_ERROR',
                    'data': {'errors': err.messages}
                }), 400
        
        return decorated_function
    return decorator


# ============= 统计分析工具函数 =============

def get_department_workload_statistics(department: str, start_date: date, end_date: date) -> Dict:
    """
    获取科室工作量统计
    
    Args:
        department: 科室名称
        start_date: 开始日期
        end_date: 结束日期
    
    Returns:
        Dict: 统计数据
    """
    doctors = Doctor.query.filter_by(department=department, status='active').all()
    
    total_appointments = 0
    total_schedules = 0
    
    for doctor in doctors:
        try:
            appointments = Appointment.query.filter(
                Appointment.doctor_id == doctor.id,
                Appointment.appointment_date >= start_date,
                Appointment.appointment_date <= end_date
            ).count()
            total_appointments += appointments
        except:
            pass
        
        try:
            schedules = DoctorSchedule.query.filter(
                DoctorSchedule.doctor_id == doctor.id,
                DoctorSchedule.date >= start_date,
                DoctorSchedule.date <= end_date
            ).count()
            total_schedules += schedules
        except:
            pass
    
    return {
        'department': department,
        'doctor_count': len(doctors),
        'total_appointments': total_appointments,
        'total_schedules': total_schedules,
        'avg_appointments_per_doctor': total_appointments / len(doctors) if doctors else 0
    }


def log_operation_decorator(operation: str, resource: str):
    """
    操作日志装饰器
    
    Args:
        operation: 操作类型 (create/update/delete/view)
        resource: 资源类型 (doctor/schedule/performance)
    
    Returns:
        装饰器函数
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # 执行路由函数
            try:
                result = f(*args, **kwargs)
                
                # 尝试获取资源ID和名称
                resource_id = kwargs.get('doctor_id') or kwargs.get('schedule_id') or kwargs.get('performance_id')
                resource_name = None
                
                # 记录成功日志
                log_operation(
                    operation=operation,
                    resource=resource,
                    resource_id=resource_id,
                    resource_name=resource_name,
                    status='success'
                )
                
                return result
            except Exception as e:
                # 记录失败日志
                resource_id = kwargs.get('doctor_id') or kwargs.get('schedule_id') or kwargs.get('performance_id')
                log_operation(
                    operation=operation,
                    resource=resource,
                    resource_id=resource_id,
                    status='failed',
                    error_message=str(e)
                )
                raise
        
        return decorated_function
    return decorator


def check_schedule_conflict_enhanced(doctor_id: int, schedule_date: date, shift: str,
                                     start_time: Optional[str] = None, end_time: Optional[str] = None,
                                     exclude_schedule_id: Optional[int] = None) -> tuple[bool, Optional[str]]:
    """
    增强的排班冲突检测（支持班次和时间段检测）
    
    Args:
        doctor_id: 医生ID
        schedule_date: 排班日期
        shift: 班次
        start_time: 开始时间（可选）
        end_time: 结束时间（可选）
        exclude_schedule_id: 排除的排班ID
    
    Returns:
        tuple: (是否有冲突, 冲突信息)
    """
    query = DoctorSchedule.query.filter_by(
        doctor_id=doctor_id,
        date=schedule_date
    )
    
    if exclude_schedule_id:
        query = query.filter(DoctorSchedule.id != exclude_schedule_id)
    
    # 检查同一天是否已有该班次
    existing_shift = query.filter_by(shift=shift).first()
    if existing_shift:
        return True, f"医生在{schedule_date}的{shift}班次已存在排班"
    
    # 如果提供了时间段，检查时间冲突
    if start_time and end_time:
        conflicts = query.filter(
            DoctorSchedule.start_time.isnot(None),
            DoctorSchedule.end_time.isnot(None),
            db.or_(
                db.and_(
                    DoctorSchedule.start_time <= start_time,
                    DoctorSchedule.end_time > start_time
                ),
                db.and_(
                    DoctorSchedule.start_time < end_time,
                    DoctorSchedule.end_time >= end_time
                ),
                db.and_(
                    DoctorSchedule.start_time >= start_time,
                    DoctorSchedule.end_time <= end_time
                )
            )
        ).first()
        
        if conflicts:
            return True, f"医生在{schedule_date} {start_time}-{end_time}时间段存在排班冲突"
    
    # 检查医生是否有请假记录
    try:
        from modules.doctor.models_extended import DoctorLeave
        leave = DoctorLeave.query.filter(
            DoctorLeave.doctor_id == doctor_id,
            DoctorLeave.status == 'approved',
            DoctorLeave.start_date <= schedule_date,
            DoctorLeave.end_date >= schedule_date
        ).first()
        
        if leave:
            return True, f"医生在{schedule_date}有已批准的请假记录"
    except:
        pass
    
    return False, None
